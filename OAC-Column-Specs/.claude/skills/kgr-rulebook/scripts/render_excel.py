# -*- coding: utf-8 -*-
"""
Bước GỘP CUỐI: nhiều rulebook JSON (schema gọn) → 1 file Excel bàn giao.
Sheet 0 = Glossary; mỗi báo cáo 1 sheet (summary: 2 khối cột/dòng); cột confirm Y/N (dropdown).
Dùng: python render_excel.py <glossary.json> <out.xlsx> <rulebook1.json> [rulebook2.json ...]
Rulebook JSON: {workbook, canvas, viz_title, is_summary, rows:[{block:"column|metric", name, calc, exclusions, note}]}
"""
import json, sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

glo_path, out = sys.argv[1], sys.argv[2]
rbs = sys.argv[3:]
GREEN = "44BA46"
HF = PatternFill("solid", fgColor=GREEN); HFONT = Font(bold=True, color="FFFFFF", size=10)
SUB = PatternFill("solid", fgColor="2E7D32"); SUBF = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top"); CEN = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin = Side(style="thin", color="DDDDDD"); BD = Border(thin, thin, thin, thin)
COLS_BASE = ["Cột / Chỉ tiêu", "Cách tính (logic nghiệp vụ)", "Loại trừ / Bộ lọc", "Ghi chú", "KGR xác nhận"]
W_BASE = [26, 70, 40, 34, 12]
COLS_QA = ["Cột / Chỉ tiêu", "Cách tính (logic nghiệp vụ)", "Loại trừ / Bộ lọc", "Ghi chú",
           "Cách tính (QA)", "Loại trừ-Bộ lọc (QA)", "Ghi chú (QA)", "KGR xác nhận"]
W_QA = [24, 48, 30, 24, 48, 30, 24, 12]
COLS, W, QA = COLS_BASE, W_BASE, False

def safe(s): return re.sub(r'[\\/*?:\[\]]', '_', s or "sheet")[:31]

wb = Workbook(); ws0 = wb.active; ws0.title = "0. Glossary"
ws0.column_dimensions["A"].width = 30; ws0.column_dimensions["B"].width = 95
ws0["A1"] = "RULE-BOOK BÁO CÁO KGR — Thuật ngữ dùng chung (xác nhận 1 lần)"; ws0["A1"].font = Font(bold=True, size=13, color=GREEN)
glo = json.load(open(glo_path, encoding="utf-8")); r = 3
for c, h in ((1, "Thuật ngữ"), (2, "Định nghĩa")):
    ws0.cell(row=r, column=c, value=h).fill = HF; ws0.cell(row=r, column=c).font = HFONT
r += 1
for t in glo["terms"]:
    ws0.cell(row=r, column=1, value=t["term"]).alignment = WRAP
    ws0.cell(row=r, column=2, value=t["definition"]).alignment = WRAP
    r += 1
ws0.freeze_panes = "A4"

def hdr(ws, row):
    for i, h in enumerate(COLS):
        c = ws.cell(row=row, column=i + 1, value=h); c.fill = HF; c.font = HFONT; c.alignment = CEN; c.border = BD
    ws.row_dimensions[row].height = 28

for rbp in rbs:
    rb = json.load(open(rbp, encoding="utf-8")); ws = wb.create_sheet(safe(rb.get("canvas")))
    QA = bool(rb.get("qa_phase"))
    COLS, W = (COLS_QA, W_QA) if QA else (COLS_BASE, W_BASE)
    for i, w in enumerate(W):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.cell(row=1, column=1, value=f"{rb['workbook']} · {rb.get('canvas')} · {rb.get('viz_title','')}").font = Font(bold=True, size=12, color=GREEN)
    r = 3
    def block(title, rows):
        global r
        if not rows: return
        if title:
            c = ws.cell(row=r, column=1, value=title); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            c.fill = SUB; c.font = SUBF; r += 1
        hdr(ws, r); r += 1
        for it in rows:
            base = [it["name"], it.get("calc", ""), it.get("exclusions", "—"), it.get("note", "")]
            qa = [it.get("qa1_calc", ""), it.get("qa1_exclusions", ""), it.get("qa1_note", "")] if QA else []
            for i, v in enumerate(base + qa + [""]):
                cell = ws.cell(row=r, column=i + 1, value=v); cell.alignment = WRAP; cell.border = BD
                if i == 0: cell.font = Font(bold=True)
            r += 1
    if rb.get("is_summary"):
        block("A. Các CỘT giá trị", [x for x in rb["rows"] if x["block"] == "column"])
        block("B. Các DÒNG chỉ tiêu", [x for x in rb["rows"] if x["block"] == "metric"])
    else:
        secs = {}
        for it in rb["rows"]:
            secs.setdefault(it.get("section"), []).append(it)
        for sec, rws in secs.items():
            block(sec, rws)
    cl = get_column_letter(len(COLS)); dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True); ws.add_data_validation(dv); dv.add(f"{cl}4:{cl}{r}")
    ws.freeze_panes = "A3"

wb.save(out); print("wrote", out, "| sheets:", [w.title for w in wb.worksheets])
